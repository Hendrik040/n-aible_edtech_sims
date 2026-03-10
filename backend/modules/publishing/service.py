"""
Publishing service for business logic.

Handles all business logic for the publishing module including
simulation saving, publishing, and file storage operations.
"""

import asyncio
import base64
import logging
from typing import List, Optional, Dict, Any, Tuple
from datetime import datetime
from sqlalchemy.orm import Session
from fastapi import HTTPException

from common.db.models import (
    Simulation, SimulationPersona, SimulationScene, SimulationFile,
    User, scene_personas
)
from common.services.s3_service import s3_service
from common.utils.id_generator import generate_simulation_id
from .repository import PublishingRepository
from .tasks import (
    is_temporary_image_url as _is_temporary_image_url,
    is_s3_url as _is_s3_url,
    handle_image_uploads,
    enqueue_image_upload,
    get_upload_status,
    check_image_exists_in_s3
)

logger = logging.getLogger(__name__)


class PublishingService:
    """Service for publishing business logic."""
    
    def __init__(self, db: Session):
        self.db = db
        self.repository = PublishingRepository(db)
    
    async def handle_pdf_storage(
        self,
        simulation: Simulation,
        pdf_metadata: Dict[str, Any]
    ) -> None:
        """Handle PDF storage to S3."""
        logger.info(f"[PDF_STORAGE] Starting PDF storage for simulation {simulation.id}")
        
        filename = pdf_metadata.get("filename")
        if not filename:
            logger.warning("[PDF_STORAGE] Missing filename in PDF metadata")
            return
        
        file_size = pdf_metadata.get("file_size")
        file_type = pdf_metadata.get("file_type", "application/pdf")
        wasabi_url = pdf_metadata.get("wasabi_url")
        pdf_url = pdf_metadata.get("pdf_url")
        file_contents_base64 = pdf_metadata.get("file_contents_base64")
        temp_pdf_url = pdf_metadata.get("temp_pdf_url")
        
        # Check if PDF already exists in database
        existing_file = self.repository.get_simulation_file(simulation.id, filename)
        existing_case_study_url = getattr(simulation, 'case_study_url', None)
        
        if existing_case_study_url and _is_s3_url(existing_case_study_url):
            logger.info(f"[PDF_STORAGE] Simulation already has case_study_url: {existing_case_study_url}")
            return
        
        if existing_file and existing_file.file_path and _is_s3_url(existing_file.file_path):
            logger.info(f"[PDF_STORAGE] PDF already exists in database with S3 URL")
            if not existing_case_study_url:
                setattr(simulation, 'case_study_url', existing_file.file_path)
                self.db.add(simulation)
            return
        
        # Check if PDF already exists in S3
        s3_key = s3_service.get_case_study_key(simulation.id, filename)
        if await s3_service.file_exists(s3_key):
            s3_url = s3_service._build_public_url(s3_key)
            logger.info(f"[PDF_STORAGE] PDF already exists in S3: {s3_key}")
            
            self.repository.create_or_update_simulation_file(
                simulation.id, filename, s3_url, file_size, file_type
            )
            setattr(simulation, 'case_study_url', s3_url)
            self.db.add(simulation)
            return
        
        # Check if metadata already has S3 URL
        existing_url = None
        if wasabi_url and _is_s3_url(wasabi_url):
            existing_url = wasabi_url
        elif pdf_url and _is_s3_url(pdf_url):
            existing_url = pdf_url
        
        if existing_url:
            self.repository.create_or_update_simulation_file(
                simulation.id, filename, existing_url, file_size, file_type
            )
            setattr(simulation, 'case_study_url', existing_url)
            self.db.add(simulation)
            return
        
        # Upload from base64
        if file_contents_base64:
            pdf_bytes = base64.b64decode(file_contents_base64)
            s3_url = await s3_service.upload_from_bytes(pdf_bytes, s3_key, file_type)
            
            if s3_url:
                self.repository.create_or_update_simulation_file(
                    simulation.id, filename, s3_url, file_size, file_type
                )
                setattr(simulation, 'case_study_url', s3_url)
                self.db.add(simulation)
                logger.info(f"[PDF_STORAGE] Uploaded PDF to S3: {s3_url}")
            return
        
        # Upload from temp URL
        if temp_pdf_url:
            # Download from temp location and upload to final location
            # Extract S3 key from temp URL if it's already in S3
            if _is_s3_url(temp_pdf_url):
                # Extract key from URL
                parts = temp_pdf_url.split('/')
                if 'scenarios' in temp_pdf_url:
                    # Find the scenarios part and extract the key
                    idx = temp_pdf_url.find('scenarios/')
                    temp_key = temp_pdf_url[idx:]
                    pdf_bytes = await s3_service.download_file(temp_key)
                    if pdf_bytes:
                        s3_url = await s3_service.upload_from_bytes(pdf_bytes, s3_key, file_type)
                        if s3_url:
                            self.repository.create_or_update_simulation_file(
                                simulation.id, filename, s3_url, file_size, file_type
                            )
                            setattr(simulation, 'case_study_url', s3_url)
                            self.db.add(simulation)
                            # Delete temp file
                            await s3_service.delete_file(temp_key)
                            logger.info(f"[PDF_STORAGE] Moved PDF from temp to final location: {s3_url}")
            else:
                # Download from external URL and upload
                s3_url = await s3_service.upload_from_url(temp_pdf_url, s3_key, file_type)
                if s3_url:
                    self.repository.create_or_update_simulation_file(
                        simulation.id, filename, s3_url, file_size, file_type
                    )
                    setattr(simulation, 'case_study_url', s3_url)
                    self.db.add(simulation)
                    logger.info(f"[PDF_STORAGE] Uploaded PDF from URL to S3: {s3_url}")
    
    async def handle_image_uploads(
        self,
        personas_to_upload: List[Dict[str, Any]],
        scenes_to_upload: List[Dict[str, Any]]
    ) -> Tuple[int, int]:
        """Handle image uploads to S3 - delegates to tasks module."""
        return await handle_image_uploads(self.db, self.repository, personas_to_upload, scenes_to_upload)
    
    async def save_simulation_draft(
        self,
        simulation_id: Optional[int],
        user_id: Optional[int],
        data: Dict[str, Any]
    ) -> Simulation:
        """Save or update a simulation draft with personas and scenes."""
        logger.info(f"[SAVE] Saving simulation draft - simulation_id: {simulation_id}, user_id: {user_id}")
        
        # Get or create simulation
        if simulation_id:
            simulation = self.repository.get_simulation_by_id(simulation_id)
            if not simulation:
                raise HTTPException(status_code=404, detail="Simulation not found")
            if user_id and simulation.created_by != user_id:
                raise HTTPException(status_code=403, detail="You can only edit simulations you created")
        else:
            # Create new simulation
            unique_id = generate_simulation_id(self.db)
            simulation = Simulation(
                unique_id=unique_id,
                title=data.get("title", "Untitled Simulation"),
                description=data.get("description", ""),
                created_by=user_id,
                status="draft",
                is_draft=True
            )
            self.db.add(simulation)
            self.db.flush()  # Get the ID
        
        # Update simulation fields
        if "title" in data:
            simulation.title = data["title"]
        if "description" in data:
            simulation.description = data["description"]
        if "student_role" in data:
            simulation.student_role = data["student_role"]
        if "learning_outcomes" in data:
            learning_outcomes = data["learning_outcomes"]
            if isinstance(learning_outcomes, str):
                # Convert string to list
                learning_outcomes = [item.strip() for item in learning_outcomes.split('\n') if item.strip()]
            simulation.learning_objectives = learning_outcomes
        
        # Update completion status
        completion_status = data.get("completion_status", {})
        if completion_status:
            simulation.name_completed = completion_status.get("name_completed", False)
            simulation.description_completed = completion_status.get("description_completed", False)
            simulation.student_role_completed = completion_status.get("student_role_completed", False)
            simulation.personas_completed = completion_status.get("personas_completed", False)
            simulation.scenes_completed = completion_status.get("scenes_completed", False)
            simulation.images_completed = completion_status.get("images_completed", False)
            simulation.learning_outcomes_completed = completion_status.get("learning_outcomes_completed", False)
            simulation.ai_enhancement_completed = completion_status.get("ai_enhancement_completed", False)
        
        # Save grading config
        if "grading_prompt" in data:
            simulation.grading_prompt = data["grading_prompt"]
        if "rubric_title" in data or "rubric_criteria" in data or "rubric_performance_levels" in data:
            grading_config = {}
            if "rubric_title" in data:
                grading_config["title"] = data["rubric_title"]
            if "rubric_criteria" in data:
                grading_config["criteria"] = data["rubric_criteria"]
            if "rubric_performance_levels" in data:
                grading_config["performance_levels"] = data["rubric_performance_levels"]
            simulation.grading_config = grading_config
        
        simulation.updated_at = datetime.utcnow()
        self.db.add(simulation)
        self.db.flush()
        
        # Save personas
        if "personas" in data and isinstance(data["personas"], list):
            # Get IDs of personas in the new data, but ONLY valid integer IDs
            # Frontend sends temp string IDs like "persona-1766273215884-0" which should NOT
            # trigger soft-delete logic. Only real database integer IDs should.
            new_persona_ids = set()
            for persona_data in data["personas"]:
                persona_id = persona_data.get("id")
                if persona_id:
                    # Only include valid integer IDs (real DB IDs)
                    if isinstance(persona_id, int):
                        new_persona_ids.add(persona_id)
                    elif isinstance(persona_id, str) and persona_id.isdigit():
                        new_persona_ids.add(int(persona_id))
                    # Skip temp IDs like "persona-1766273215884-0"

            existing_personas = self.repository.get_simulation_personas(simulation.id)

            # Only perform ID-based soft delete if we actually received VALID integer persona IDs.
            # Frontend often sends temp string IDs (e.g., "persona-1766273215884-0") which should
            # NOT trigger soft-delete. We only soft-delete when we have real DB IDs to compare.
            if new_persona_ids:
                # Soft delete personas that are not in the new data
                for persona in existing_personas:
                    if persona.id not in new_persona_ids:
                        persona.deleted_at = datetime.utcnow()
            else:
                logger.debug(
                    "[SAVE] No persona IDs provided in payload; skipping persona soft-delete "
                    "and relying on name-based matching to update/create personas."
                )

            # Create/update personas
            for idx, persona_data in enumerate(data["personas"]):
                persona_id = persona_data.get("id")
                persona = None
                
                if persona_id:
                    # Only query if persona_id is a valid integer (not a temporary string ID from frontend)
                    # Frontend sends temporary IDs like "persona-1765519956424-0" for new personas
                    is_valid_int = False
                    if isinstance(persona_id, int):
                        is_valid_int = True
                    elif isinstance(persona_id, str) and persona_id.isdigit():
                        is_valid_int = True
                        persona_id = int(persona_id)  # Convert to int for query
                    
                    if is_valid_int:
                        # Try to find existing non-deleted persona
                        persona = self.db.query(SimulationPersona).filter(
                            SimulationPersona.id == persona_id,
                            SimulationPersona.simulation_id == simulation.id,
                            SimulationPersona.deleted_at.is_(None)
                        ).first()
                        if persona:
                            logger.debug(f"[SAVE] Found persona by ID {persona_id}: '{persona.name}'")
                    else:
                        logger.debug(f"[SAVE] Persona ID '{persona_id}' is not a valid integer (likely temp ID from frontend), will try name lookup")
                
                # Handle image URL - never save temp URLs, only S3 URLs
                image_url = persona_data.get("imageUrl")
                if image_url and _is_temporary_image_url(image_url):
                    # Don't save temp URL - will be uploaded by worker
                    image_url = None
                elif image_url and not _is_s3_url(image_url):
                    # If not S3 URL and not temp, assume it's invalid
                    image_url = None
                
                if not persona:
                    # ID lookup failed - try to find by name to avoid recreating personas with new IDs
                    # This prevents losing S3 image references when personas are recreated
                    persona_name = persona_data.get("name")
                    if persona_name:
                        persona = self.db.query(SimulationPersona).filter(
                            SimulationPersona.name == persona_name,
                            SimulationPersona.simulation_id == simulation.id,
                            SimulationPersona.deleted_at.is_(None)
                        ).first()
                    
                    if not persona:
                        # Create new persona (either no ID provided, or ID/name lookup both failed)
                        # Set all fields before flushing to avoid NOT NULL constraint errors
                        persona = SimulationPersona(
                            simulation_id=simulation.id,
                            name=persona_data.get("name", f"Persona {idx + 1}"),
                            role=persona_data.get("role", ""),
                            background=persona_data.get("background"),
                            current_context=persona_data.get("current_context"),
                            correlation=persona_data.get("correlation"),
                            primary_goals=persona_data.get("primary_goals"),
                            personality_traits=persona_data.get("personality_traits"),
                            knowledge_areas=persona_data.get("knowledge_areas"),
                            communication_style=persona_data.get("communication_style"),
                            system_prompt=persona_data.get("systemPrompt"),
                            image_url=image_url  # Only S3 URLs or None
                        )
                        self.db.add(persona)
                        self.db.flush()
                        logger.info(f"[SAVE] ✅ CREATED new persona '{persona.name}' (ID: {persona.id}) for simulation {simulation.id}")
                        
                        # Temp URLs will be handled after commit via handle_image_uploads
                    else:
                        # Found by name - update it (preserves ID and S3 image references)
                        logger.info(f"[SAVE] 🔄 UPDATING persona '{persona_name}' (ID: {persona.id}) - found by name match (ID lookup failed)")
                        # Update fields (same as else block below)
                        if "name" in persona_data:
                            persona.name = persona_data["name"]
                        if "role" in persona_data:
                            persona.role = persona_data["role"]
                        if "background" in persona_data:
                            persona.background = persona_data["background"]
                        if "current_context" in persona_data:
                            persona.current_context = persona_data["current_context"]
                        if "correlation" in persona_data:
                            persona.correlation = persona_data["correlation"]
                        if "primary_goals" in persona_data:
                            persona.primary_goals = persona_data["primary_goals"]
                        if "personality_traits" in persona_data:
                            persona.personality_traits = persona_data["personality_traits"]
                        if "knowledge_areas" in persona_data:
                            persona.knowledge_areas = persona_data["knowledge_areas"]
                        if "communication_style" in persona_data:
                            persona.communication_style = persona_data["communication_style"]
                        if "systemPrompt" in persona_data:
                            persona.system_prompt = persona_data["systemPrompt"]
                        if "imageUrl" in persona_data:
                            if _is_temporary_image_url(persona_data.get("imageUrl")):
                                persona.image_url = None
                            elif _is_s3_url(persona_data.get("imageUrl")):
                                persona.image_url = persona_data.get("imageUrl")
                            else:
                                persona.image_url = None
                else:
                    # Update existing persona fields (found by ID)
                    logger.info(f"[SAVE] 🔄 UPDATING persona '{persona.name}' (ID: {persona.id}) - found by ID")
                    if "name" in persona_data:
                        persona.name = persona_data["name"]
                    if "role" in persona_data:
                        persona.role = persona_data["role"]
                    if "background" in persona_data:
                        persona.background = persona_data["background"]
                    if "current_context" in persona_data:
                        persona.current_context = persona_data["current_context"]
                    if "correlation" in persona_data:
                        persona.correlation = persona_data["correlation"]
                    if "primary_goals" in persona_data:
                        persona.primary_goals = persona_data["primary_goals"]
                    if "personality_traits" in persona_data:
                        persona.personality_traits = persona_data["personality_traits"]
                    if "knowledge_areas" in persona_data:
                        persona.knowledge_areas = persona_data["knowledge_areas"]
                    if "communication_style" in persona_data:
                        persona.communication_style = persona_data["communication_style"]
                    if "systemPrompt" in persona_data:
                        persona.system_prompt = persona_data["systemPrompt"]
                    if "imageUrl" in persona_data:
                        if _is_temporary_image_url(persona_data.get("imageUrl")):
                            persona.image_url = None
                        elif _is_s3_url(persona_data.get("imageUrl")):
                            persona.image_url = persona_data.get("imageUrl")
                        else:
                            persona.image_url = None
                
                persona.updated_at = datetime.utcnow()
                self.db.add(persona)
        
        # Save scenes - smart update/merge approach
        if "scenes" in data and isinstance(data["scenes"], list):
            # Get existing scenes indexed by ID
            existing_scenes = self.repository.get_simulation_scenes(simulation.id)
            existing_scenes_by_id = {scene.id: scene for scene in existing_scenes}
            
            # Track which scene IDs are still in use
            incoming_scene_ids = set()
            
            for idx, scene_data in enumerate(data["scenes"]):
                scene_id = scene_data.get("id")
                scene = None
                
                # Check if this is an existing scene we should update
                if scene_id and scene_id in existing_scenes_by_id:
                    # UPDATE existing scene by ID
                    scene = existing_scenes_by_id[scene_id]
                    incoming_scene_ids.add(scene_id)
                    logger.info(f"[SAVE] 🔄 UPDATING scene '{scene.title}' (ID: {scene_id}) - found by ID")
                else:
                    # ID lookup failed - try to find by title to avoid recreating scenes
                    # This prevents losing S3 image references when scenes are recreated with new IDs
                    scene_title = scene_data.get("title")
                    
                    if scene_title:
                        # Try to find existing scene by title (matching by title only, similar to persona name matching)
                        scene = self.db.query(SimulationScene).filter(
                            SimulationScene.simulation_id == simulation.id,
                            SimulationScene.title == scene_title,
                            SimulationScene.deleted_at.is_(None)
                        ).first()
                    
                    if scene:
                        # Found by title - update it (preserves ID and S3 image references)
                        logger.info(f"[SAVE] 🔄 UPDATING scene '{scene_title}' (ID: {scene.id}) - found by title match (ID lookup failed)")
                        incoming_scene_ids.add(scene.id)
                    else:
                        # CREATE new scene (either no ID/title provided, or lookup failed)
                        scene = SimulationScene(simulation_id=simulation.id)
                        self.db.add(scene)
                        logger.info(f"[SAVE] ✅ CREATING new scene '{scene_data.get('title', 'Untitled')}' for simulation {simulation.id}")
                
                # Update scene fields
                scene.title = scene_data.get("title", f"Scene {idx + 1}")
                # Support both scene_order and sequence_order (frontend sends sequence_order)
                # Check for None explicitly to preserve zero values
                scene_order = scene_data.get("scene_order")
                if scene_order is None:
                    scene_order = scene_data.get("sequence_order")
                if scene_order is None:
                    scene.scene_order = idx + 1
                else:
                    scene.scene_order = scene_order
                
                if "description" in scene_data:
                    scene.description = scene_data["description"]
                if "user_goal" in scene_data:
                    scene.user_goal = scene_data["user_goal"]
                if "timeout_turns" in scene_data:
                    scene.timeout_turns = scene_data["timeout_turns"]
                if "success_metric" in scene_data:
                    scene.success_metric = scene_data["success_metric"]

                # Code challenge fields
                if "scene_type" in scene_data:
                    scene.scene_type = scene_data["scene_type"]
                if "starter_code" in scene_data:
                    scene.starter_code = scene_data["starter_code"]
                if "code_grading_criteria" in scene_data:
                    scene.code_grading_criteria = scene_data["code_grading_criteria"]
                # data_files / reference_files: only persist already-uploaded entries (those
                # with s3_key) at this point. Entries that still carry raw base64 `content`
                # will be uploaded to S3 by _upload_scene_data_files() and written then.
                # This prevents raw base64 blobs from being committed if the upload fails.
                if "data_files" in scene_data and scene_data["data_files"] is not None:
                    scene.data_files = [
                        f for f in scene_data["data_files"]
                        if not (f.get("content", "")).startswith("data:")
                    ] or scene.data_files
                if "reference_files" in scene_data and scene_data["reference_files"] is not None:
                    scene.reference_files = [
                        f for f in scene_data["reference_files"]
                        if not (f.get("content", "")).startswith("data:")
                    ] or scene.reference_files

                # Handle image URL - never save temp URLs, only S3 URLs
                image_url = scene_data.get("imageUrl") or scene_data.get("image_url")
                if image_url and _is_temporary_image_url(image_url):
                    scene.image_url = None
                elif image_url and _is_s3_url(image_url):
                    scene.image_url = image_url
                # Don't clear existing image_url if no new one provided
                
                self.db.flush()  # Flush to get scene.id for new scenes
                
                # Track the scene ID (for new scenes, get it after flush)
                if scene.id:
                    incoming_scene_ids.add(scene.id)
                
                # Update scene-persona associations
                # First, delete existing associations for this scene
                self.db.execute(
                    scene_personas.delete().where(scene_personas.c.scene_id == scene.id)
                )
                
                # Then create new associations
                if "personas_involved" in scene_data and isinstance(scene_data["personas_involved"], list):
                    for persona_name in scene_data["personas_involved"]:
                        if not persona_name or not isinstance(persona_name, str):
                            continue
                        
                        persona = self.db.query(SimulationPersona).filter(
                            SimulationPersona.name == persona_name,
                            SimulationPersona.simulation_id == simulation.id,
                            SimulationPersona.deleted_at.is_(None)
                        ).first()
                        
                        if persona:
                            self.db.execute(
                                scene_personas.insert().values(
                                    scene_id=scene.id,
                                    persona_id=persona.id,
                                    involvement_level="participant"
                                )
                            )
            
            # DELETE scenes that are no longer in the incoming data
            for scene_id, scene in existing_scenes_by_id.items():
                if scene_id not in incoming_scene_ids:
                    # First delete scene_personas associations (children first!)
                    self.db.execute(
                        scene_personas.delete().where(scene_personas.c.scene_id == scene_id)
                    )
                    # Then soft delete the scene
                    scene.deleted_at = datetime.utcnow()
                    self.db.add(scene)
        
        self.db.commit()
        
        # After commit, collect temp URLs and check S3 before enqueueing
        # This prevents duplicate uploads when personas/scenes are recreated with new IDs
        personas_to_upload = []
        scenes_to_upload = []
        
        # Collect personas with temp URLs from original data
        if "personas" in data and isinstance(data["personas"], list):
            saved_personas = self.repository.get_simulation_personas(simulation.id)
            persona_by_name = {p.name: p for p in saved_personas}
            
            for persona_data in data["personas"]:
                temp_url = persona_data.get("imageUrl")
                if temp_url and _is_temporary_image_url(temp_url):
                    persona_name = persona_data.get("name")
                    persona = persona_by_name.get(persona_name)
                    if persona and not persona.image_url:  # Only if image_url is None (temp URL was set to None)
                        personas_to_upload.append({
                            "persona_id": persona.id,
                            "scenario_id": simulation.id,
                            "temp_url": temp_url
                        })
        
        # Collect scenes with temp URLs from original data
        if "scenes" in data and isinstance(data["scenes"], list):
            saved_scenes = self.repository.get_simulation_scenes(simulation.id)
            # Match scenes by order (since scenes are recreated)
            for idx, scene_data in enumerate(data["scenes"]):
                if idx < len(saved_scenes):
                    scene = saved_scenes[idx]
                    temp_url = scene_data.get("imageUrl") or scene_data.get("image_url")
                    if temp_url and _is_temporary_image_url(temp_url) and not scene.image_url:
                        scenes_to_upload.append({
                            "scene_id": scene.id,
                            "scenario_id": simulation.id,
                            "temp_url": temp_url
                        })
        
        # Enqueue uploads via handle_image_uploads (which checks S3 first)
        # This ensures S3 check happens before enqueueing, preventing duplicates
        if personas_to_upload or scenes_to_upload:
            await self.handle_image_uploads(personas_to_upload, scenes_to_upload)

        # Upload data files and reference files to S3 (for code_challenge scenes)
        await self._upload_scene_data_files(simulation.id, data)

        logger.info(f"[SAVE] Successfully saved simulation {simulation.id}")
        return simulation
    
    @staticmethod
    def _generate_data_preview(filename: str, file_bytes: bytes) -> str:
        """Generate a CSV-style preview string from a data file (CSV, XLSX, JSON)."""
        try:
            lower = filename.lower()
            if lower.endswith(".csv"):
                text = file_bytes.decode("utf-8", errors="replace")
                lines = text.split("\n")[:6]
                return "\n".join(lines)
            elif lower.endswith((".xlsx", ".xls")):
                from io import BytesIO
                from openpyxl import load_workbook
                wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
                ws = wb.active
                rows = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 6:  # header + 5 data rows
                        break
                    rows.append(",".join(str(c) if c is not None else "" for c in row))
                wb.close()
                return "\n".join(rows)
            elif lower.endswith(".json"):
                import json as _json
                data = _json.loads(file_bytes.decode("utf-8", errors="replace"))
                if isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
                    headers = list(data[0].keys())
                    lines = [",".join(headers)]
                    for row in data[:5]:
                        lines.append(",".join(str(row.get(h, "")) for h in headers))
                    return "\n".join(lines)
        except Exception as e:
            logger.warning(f"[SAVE] Preview generation failed for {filename}: {e}")
        return ""

    async def _upload_scene_data_files(
        self,
        simulation_id: int,
        data: Dict[str, Any]
    ) -> None:
        """Upload base64 data files and reference files to S3, then update scene records."""
        from common.services.s3_service import s3_service, parse_generic_data_url

        if "scenes" not in data or not isinstance(data["scenes"], list):
            return

        saved_scenes = self.repository.get_simulation_scenes(simulation_id)
        # Build lookup by scene ID so ordering differences don't misalign files
        saved_scenes_by_id = {s.id: s for s in saved_scenes}

        for idx, scene_data in enumerate(data["scenes"]):
            scene_id = scene_data.get("id")
            if scene_id and scene_id in saved_scenes_by_id:
                # Existing scene — match by stable ID
                scene = saved_scenes_by_id[scene_id]
            else:
                # No stable ID available (new scene not yet flushed, or ID missing).
                # Skip rather than falling back to index-based matching, which can
                # misroute uploads when the client sends scenes in a different order
                # than scene_order (e.g. after reordering or inserting new scenes).
                logger.warning(
                    f"[SAVE] scene at index {idx} (id={scene_id!r}) has no stable DB ID; "
                    "skipping file upload — re-save after the scene is committed"
                )
                continue

            # Process data_files
            raw_data_files = scene_data.get("data_files") or []
            updated_data_files = []
            for file_info in raw_data_files:
                content = file_info.get("content")
                filename = file_info.get("filename", "unknown")
                if content and content.startswith("data:"):
                    # New upload — push to S3
                    try:
                        file_bytes, content_type = parse_generic_data_url(content)
                        s3_key = s3_service.get_data_file_key(simulation_id, f"{scene.id}/{filename}")
                        s3_url = await s3_service.upload_from_bytes(file_bytes, s3_key, content_type)
                        if not s3_url:
                            logger.error(f"[SAVE] Upload returned no URL for data file '{filename}' in scene {scene.id}; skipping")
                            continue
                        # Generate tabular preview (headers + 5 rows)
                        preview = file_info.get("preview", "")
                        if not preview:
                            preview = self._generate_data_preview(filename, file_bytes)
                        updated_data_files.append({
                            "filename": filename,
                            "s3_key": s3_key,
                            "preview": preview,
                        })
                        logger.info(f"[SAVE] Uploaded data file '{filename}' to S3: {s3_key}")
                    except Exception as e:
                        logger.error(f"[SAVE] Failed to upload data file '{filename}': {e}")
                else:
                    # Already persisted (has s3_key) — keep as-is
                    updated_data_files.append({
                        "filename": filename,
                        "s3_key": file_info.get("s3_key", ""),
                        "preview": file_info.get("preview", ""),
                    })

            # Process reference_files
            raw_ref_files = scene_data.get("reference_files") or []
            updated_ref_files = []
            for file_info in raw_ref_files:
                content = file_info.get("content")
                filename = file_info.get("filename", "unknown")
                if content and content.startswith("data:"):
                    try:
                        file_bytes, content_type = parse_generic_data_url(content)
                        s3_key = s3_service.get_reference_file_key(simulation_id, f"{scene.id}/{filename}")
                        s3_url = await s3_service.upload_from_bytes(file_bytes, s3_key, content_type)
                        if not s3_url:
                            logger.error(f"[SAVE] Upload returned no URL for reference file '{filename}' in scene {scene.id}; skipping")
                            continue
                        updated_ref_files.append({
                            "filename": filename,
                            "s3_key": s3_key,
                        })
                        logger.info(f"[SAVE] Uploaded reference file '{filename}' to S3: {s3_key}")
                    except Exception as e:
                        logger.error(f"[SAVE] Failed to upload reference file '{filename}': {e}")
                else:
                    updated_ref_files.append({
                        "filename": filename,
                        "s3_key": file_info.get("s3_key", ""),
                    })

            # Update scene record with S3 keys (strip base64 content)
            if updated_data_files:
                scene.data_files = updated_data_files
            if updated_ref_files:
                scene.reference_files = updated_ref_files

        self.db.commit()

    async def publish_simulation(
        self,
        simulation_id: int
    ) -> Simulation:
        """Publish a simulation (make it available for assignment)."""
        simulation = self.repository.get_simulation_by_id(simulation_id)
        if not simulation:
            raise HTTPException(status_code=404, detail="Simulation not found")
        
        # Check upload status before publishing
        upload_status = get_upload_status(simulation_id)
        has_pending_uploads = upload_status["status"] == "uploading" and upload_status["pending"] > 0
        
        # Verify all images are in S3 (check personas and scenes)
        personas = self.repository.get_simulation_personas(simulation_id)
        scenes = self.repository.get_simulation_scenes(simulation_id)
        
        missing_images = []
        images_updated_from_s3 = []
        
        for persona in personas:
            # Check if image_url is None (still uploading) or is a temp URL
            if not persona.image_url:
                # Redis might show pending, but check if image actually exists in S3
                # (worker may have processed but Redis update failed)
                s3_url = await check_image_exists_in_s3(simulation_id, "persona", persona.id)
                if s3_url:
                    # Image exists in S3 but DB not updated - fix it
                    persona.image_url = s3_url
                    self.db.add(persona)
                    images_updated_from_s3.append(f"Persona '{persona.name}'")
                    logger.warning(f"[PUBLISH] Persona '{persona.name}' image exists in S3 but DB was None - updated database")
                else:
                    # Image doesn't exist - worker not running or failed
                    missing_images.append(f"Persona '{persona.name}' image still uploading (not found in S3)")
            elif _is_temporary_image_url(persona.image_url):
                missing_images.append(f"Persona '{persona.name}' image not uploaded to S3")
        
        for scene in scenes:
            # Check if image_url is None (still uploading) or is a temp URL
            if not scene.image_url:
                # Redis might show pending, but check if image actually exists in S3
                s3_url = await check_image_exists_in_s3(simulation_id, "scene", scene.id)
                if s3_url:
                    # Image exists in S3 but DB not updated - fix it
                    scene.image_url = s3_url
                    self.db.add(scene)
                    images_updated_from_s3.append(f"Scene '{scene.title}'")
                    logger.warning(f"[PUBLISH] Scene '{scene.title}' image exists in S3 but DB was None - updated database")
                else:
                    # Image doesn't exist - worker not running or failed
                    missing_images.append(f"Scene '{scene.title}' image still uploading (not found in S3)")
            elif _is_temporary_image_url(scene.image_url):
                missing_images.append(f"Scene '{scene.title}' image not uploaded to S3")
        
        # Commit any database updates from S3 checks
        if images_updated_from_s3:
            self.db.commit()
            logger.info(f"[PUBLISH] Updated {len(images_updated_from_s3)} images from S3: {', '.join(images_updated_from_s3)}")
        
        # If we still have missing images, fail loudly
        if missing_images:
            error_msg = f"Cannot publish simulation: {len(missing_images)} images not ready. {', '.join(missing_images[:3])}"
            if has_pending_uploads:
                error_msg += f" (Redis shows {upload_status['pending']} pending uploads - worker may not be running)"
            raise HTTPException(
                status_code=400,
                detail=error_msg
            )
        
        # If Redis shows pending but all images are in S3, log warning
        if has_pending_uploads:
            logger.warning(f"[PUBLISH] Redis shows {upload_status['pending']} pending uploads but all images exist in S3 - Redis status may be stale")
        
        simulation.is_draft = False
        simulation.is_public = True
        simulation.status = "active"
        simulation.updated_at = datetime.utcnow()
        
        self.db.add(simulation)
        self.db.commit()
        
        logger.info(f"[PUBLISH] Published simulation {simulation_id}")
        return simulation
    
    def update_simulation_status(
        self,
        simulation_id: int,
        status: str,
        user_id: Optional[int] = None
    ) -> Simulation:
        """Update simulation status (draft, active, archived, creating)."""
        logger.info(f"[STATUS_UPDATE] Updating simulation {simulation_id} to status {status}")
        
        simulation = self.repository.get_simulation_by_id(simulation_id)
        if not simulation:
            raise HTTPException(status_code=404, detail="Simulation not found")
        
        # Check permissions - user can only update their own simulations
        if user_id and simulation.created_by != user_id:
            raise HTTPException(
                status_code=403,
                detail="You can only update simulations you created"
            )
        
        # Validate status
        valid_statuses = ["draft", "active", "archived", "creating"]
        if status not in valid_statuses:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid status. Must be one of: {', '.join(valid_statuses)}"
            )
        
        # Update status and related fields
        simulation.status = status
        
        if status == "active":
            # Publishing: make it available for assignment
            simulation.is_draft = False
            simulation.is_public = True
        elif status == "draft":
            # Unpublishing: make it unavailable
            simulation.is_draft = True
            simulation.is_public = False
        # For "archived" and "creating", keep existing is_draft and is_public values
        
        simulation.updated_at = datetime.utcnow()
        
        self.db.add(simulation)
        self.db.commit()
        
        logger.info(f"[STATUS_UPDATE] Successfully updated simulation {simulation_id} to {status}")
        return simulation
    
    def delete_simulation(
        self,
        simulation_id: int,
        user_id: Optional[int] = None
    ) -> bool:
        """Delete a simulation (soft delete)."""
        logger.info(f"[DELETE] Deleting simulation {simulation_id}, user_id: {user_id}")
        
        success = self.repository.delete_simulation(simulation_id, user_id)
        if not success:
            if not self.repository.get_simulation_by_id(simulation_id):
                raise HTTPException(status_code=404, detail="Simulation not found")
            else:
                raise HTTPException(status_code=403, detail="You can only delete simulations you created")
        
        logger.info(f"[DELETE] Successfully deleted simulation {simulation_id}")
        return True
